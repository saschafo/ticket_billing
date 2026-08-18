"""Kennzahlen für die Geschäftsführung.

Wie in ``dashboard.py`` gilt: Die Abfragen umgehen die Zeilenfilter, weil
eine Auswertung zählen muss, was der Aufrufer einzeln nicht sehen darf.
Genau deshalb steht vor jeder Abfrage die ausdrückliche Prüfung, wer das
darf.

Alle Zeitangaben in Stunden, alle Abfragen auf denselben Zeitraum bezogen --
sonst vergleicht man Zahlen, die nicht zusammengehören.
"""

import frappe
from frappe import _
from frappe.utils import add_days, flt, nowdate

from ticket_billing.constants import (
	CLOSED_STATUSES,
	FIELD_ASSIGNEE,
	FIELD_DEPARTMENT,
	FIELD_FIRST_RESPONSE,
	FIELD_ORIGIN,
	FIELD_RESOLVED,
	OPEN_STATUSES,
	ORIGIN_EXTERNAL,
)
from ticket_billing.utils.context import get_scope_departments, is_lead, is_unrestricted

#: 84 Tage sind genau zwölf Wochen -- die Voreinstellung der
#: Abteilungssicht, damit das Liniendiagramm zwölf volle Punkte zeigt.
ALLOWED_PERIODS = (7, 30, 84, 90, 365)


def _require_management() -> None:
	if not is_unrestricted():
		frappe.throw(
			_("This evaluation is restricted to management."), frappe.PermissionError
		)


def _period(days) -> tuple[int, str]:
	days = frappe.utils.cint(days) or 30
	if days not in ALLOWED_PERIODS:
		days = 30
	return days, add_days(nowdate(), -days)


def _short(name: str) -> str:
	"""Firmenkürzel abschneiden -- "Support - MF" wird zu "Support"."""
	return str(name or "").split(" - ")[0]


# ---------------------------------------------------------------------------
# Bausteine
# ---------------------------------------------------------------------------


def _department_tickets(since: str) -> dict[str, dict]:
	"""Aufkommen, Herkunft, Reaktions- und Lösungszeit je Abteilung.

	Bezugsgröße ist das **Anlagedatum**: Gezählt werden Tickets, die im
	Zeitraum entstanden sind. Sonst mischten sich alte Vorgänge in die
	Reaktionszeit, und der Vergleich zweier Zeiträume wäre wertlos.
	"""
	rows = frappe.db.sql(
		f"""
		select
			`{FIELD_DEPARTMENT}` as department,
			count(name) as total,
			sum(case when status in %(open_statuses)s then 1 else 0 end) as open_count,
			sum(case when status in %(closed_statuses)s then 1 else 0 end) as closed_count,
			sum(case when `{FIELD_ORIGIN}` = %(external)s then 1 else 0 end) as external_count,
			sum(case when `{FIELD_ORIGIN}` != %(external)s then 1 else 0 end) as internal_count,
			avg(case
				when `{FIELD_FIRST_RESPONSE}` is not null
				then timestampdiff(second, creation, `{FIELD_FIRST_RESPONSE}`) / 3600
			end) as avg_response_hours,
			avg(case
				when `{FIELD_RESOLVED}` is not null
				then timestampdiff(second, creation, `{FIELD_RESOLVED}`) / 3600
			end) as avg_resolution_hours
		from `tabIssue`
		where creation >= %(since)s and `{FIELD_DEPARTMENT}` is not null
		group by `{FIELD_DEPARTMENT}`
		""",
		{
			"since": since,
			"open_statuses": list(OPEN_STATUSES),
			"closed_statuses": list(CLOSED_STATUSES),
			"external": ORIGIN_EXTERNAL,
		},
		as_dict=True,
	)
	return {r.department: r for r in rows}


def _department_hours(since: str) -> dict[str, dict]:
	"""Stunden je Abteilung, getrennt nach abrechenbar und intern.

	Abrechenbar heißt hier: Der Zeiteintrag hängt an einem Ticket externer
	Herkunft. Die Unterscheidung kommt also vom Vorgang, nicht von einem
	Kennzeichen am Zeiteintrag -- so kann sie nicht auseinanderlaufen.
	"""
	rows = frappe.db.sql(
		f"""
		select
			ts.department as department,
			sum(td.hours) as hours_total,
			sum(case when i.`{FIELD_ORIGIN}` = %(external)s then td.hours else 0 end) as hours_billable,
			sum(case when i.`{FIELD_ORIGIN}` != %(external)s then td.hours else 0 end) as hours_internal,
			sum(case when ts.docstatus = 1 then td.hours else 0 end) as hours_submitted,
			sum(case when ts.docstatus = 0 then td.hours else 0 end) as hours_draft
		from `tabTimesheet Detail` td
		inner join `tabTimesheet` ts on ts.name = td.parent
		left join `tabIssue` i on i.name = td.`tb_issue`
		where ts.docstatus < 2 and td.from_time >= %(since)s and ts.department is not null
		group by ts.department
		""",
		{"since": since, "external": ORIGIN_EXTERNAL},
		as_dict=True,
	)
	return {r.department: r for r in rows}


def _trend(since: str, daily: bool) -> list[dict]:
	"""Ticketaufkommen im Zeitverlauf, aufgeschlüsselt nach Abteilung."""
	bucket = "date(creation)" if daily else "date_format(creation, '%%x-KW%%v')"

	rows = frappe.db.sql(
		f"""
		select {bucket} as period, `{FIELD_DEPARTMENT}` as department,
		       count(name) as count, min(creation) as sort_key
		from `tabIssue`
		where creation >= %(since)s and `{FIELD_DEPARTMENT}` is not null
		group by period, department
		order by sort_key asc
		""",
		{"since": since},
		as_dict=True,
	)

	counted: dict[str, dict] = {}
	for row in rows:
		key = str(row.period)
		entry = counted.setdefault(key, {"period": key, "values": {}, "total": 0})
		entry["values"][_short(row.department)] = row.count
		entry["total"] += row.count

	# Leere Zeiträume auffüllen. Ohne das zeigt die Achse nur Tage, an denen
	# etwas passiert ist -- eine Woche Pause sähe dann aus wie zwei
	# aufeinanderfolgende Tage, und "im Zeitverlauf" wäre das Gegenteil von
	# dem, was die Grafik behauptet.
	return [
		counted.get(key, {"period": key, "values": {}, "total": 0})
		for key in _period_keys(since, daily)
	]


def _period_keys(since: str, daily: bool) -> list[str]:
	"""Alle Zeitpunkte des Zeitraums, in der Schreibweise der SQL-Gruppierung."""
	from datetime import timedelta

	start = frappe.utils.getdate(since)
	end = frappe.utils.getdate(nowdate())

	if daily:
		days = (end - start).days
		return [str(start + timedelta(days=i)) for i in range(days + 1)]

	# Tageweise durchgehen und Wochen einsammeln. Höchstens 365 Schritte, dafür
	# ohne die Sonderfälle, die ein Sprung um sieben Tage am Jahreswechsel
	# oder beim ersten unvollständigen Woche mit sich brächte.
	keys, seen = [], set()
	for offset in range((end - start).days + 1):
		iso = (start + timedelta(days=offset)).isocalendar()
		# Gleiche Schreibweise wie date_format(..., '%x-KW%v'): ISO-Jahr und
		# ISO-Woche, zweistellig.
		key = f"{iso[0]}-KW{iso[1]:02d}"
		if key not in seen:
			seen.add(key)
			keys.append(key)

	return keys


def _employees(since: str) -> list[dict]:
	"""Auslastung je Mitarbeiter: offene Tickets und erfasste Zeit."""
	open_rows = frappe.db.sql(
		f"""
		select `{FIELD_ASSIGNEE}` as employee, count(name) as open_tickets
		from `tabIssue`
		where `{FIELD_ASSIGNEE}` is not null and status in %(open_statuses)s
		group by `{FIELD_ASSIGNEE}`
		""",
		{"open_statuses": list(OPEN_STATUSES)},
		as_dict=True,
	)

	hour_rows = frappe.db.sql(
		"""
		select ts.employee as employee, sum(td.hours) as hours,
		       sum(case when ts.docstatus = 1 then td.hours else 0 end) as hours_submitted
		from `tabTimesheet Detail` td
		inner join `tabTimesheet` ts on ts.name = td.parent
		where ts.docstatus < 2 and td.from_time >= %(since)s
		group by ts.employee
		""",
		{"since": since},
		as_dict=True,
	)

	ids = {r.employee for r in open_rows} | {r.employee for r in hour_rows}
	if not ids:
		return []

	people = {
		e.name: e
		for e in frappe.get_all(
			"Employee",
			filters={"name": ["in", list(ids)]},
			fields=["name", "employee_name", "department"],
		)
	}
	open_by = {r.employee: r.open_tickets for r in open_rows}
	hours_by = {r.employee: r for r in hour_rows}

	result = []
	for employee_id in ids:
		person = people.get(employee_id)
		if not person:
			continue
		hours = hours_by.get(employee_id)
		result.append(
			{
				"employee": employee_id,
				"employee_name": person.employee_name or employee_id,
				"department": _short(person.department),
				"open_tickets": open_by.get(employee_id, 0),
				"hours": flt(hours.hours if hours else 0, 2),
				"hours_submitted": flt(hours.hours_submitted if hours else 0, 2),
			}
		)

	return sorted(result, key=lambda r: (-r["open_tickets"], r["employee_name"]))


# ---------------------------------------------------------------------------
# Schnittstelle
# ---------------------------------------------------------------------------


def build_kpis(days) -> dict:
	days, since = _period(days)

	tickets = _department_tickets(since)
	hours = _department_hours(since)

	departments = []
	for name in sorted(set(tickets) | set(hours)):
		t = tickets.get(name)
		h = hours.get(name)
		# int(): SUM(CASE ...) liefert in MariaDB einen Dezimalwert, und
		# "11.0 Tickets" liest sich falsch.
		def count(field):
			return int(flt(t.get(field))) if t else 0

		total = count("total")
		external = count("external_count")

		departments.append(
			{
				"department": name,
				"label": _short(name),
				"total": total,
				"open": count("open_count"),
				"closed": count("closed_count"),
				"internal": count("internal_count"),
				"external": external,
				# Anteil in Prozent, damit Abteilungen unterschiedlicher Größe
				# vergleichbar sind -- absolute Zahlen sagen darüber nichts.
				"external_share": flt(external / total * 100, 1) if total else 0,
				"avg_response_hours": flt(t.avg_response_hours, 2) if t and t.avg_response_hours else None,
				"avg_resolution_hours": flt(t.avg_resolution_hours, 2)
				if t and t.avg_resolution_hours
				else None,
				"hours_total": flt(h.hours_total if h else 0, 2),
				"hours_billable": flt(h.hours_billable if h else 0, 2),
				"hours_internal": flt(h.hours_internal if h else 0, 2),
				"hours_submitted": flt(h.hours_submitted if h else 0, 2),
				"hours_draft": flt(h.hours_draft if h else 0, 2),
			}
		)

	employees = _employees(since)

	def total_of(key):
		return flt(sum(d[key] for d in departments), 2)

	# Durchschnitt über die Abteilungen, gewichtet nach Ticketzahl -- ein
	# ungewichteter Mittelwert gäbe einer Abteilung mit zwei Vorgängen
	# dasselbe Gewicht wie einer mit zweihundert.
	def weighted(key):
		pairs = [(d[key], d["total"]) for d in departments if d[key] is not None and d["total"]]
		if not pairs:
			return None
		weight = sum(w for _, w in pairs)
		return flt(sum(v * w for v, w in pairs) / weight, 2) if weight else None

	return {
		"days": days,
		"since": since,
		"departments": departments,
		"employees": employees,
		"trend": _trend(since, daily=days <= 31),
		"trend_granularity": "day" if days <= 31 else "week",
		"totals": {
			"tickets": sum(d["total"] for d in departments),
			"open": sum(d["open"] for d in departments),
			"closed": sum(d["closed"] for d in departments),
			"external": sum(d["external"] for d in departments),
			"internal": sum(d["internal"] for d in departments),
			"hours_total": total_of("hours_total"),
			"hours_billable": total_of("hours_billable"),
			"hours_internal": total_of("hours_internal"),
			"hours_submitted": total_of("hours_submitted"),
			"hours_draft": total_of("hours_draft"),
			"avg_response_hours": weighted("avg_response_hours"),
			"avg_resolution_hours": weighted("avg_resolution_hours"),
		},
	}


@frappe.whitelist()
def get_management_kpis(days: int | str = 30):
	_require_management()
	return build_kpis(days)


@frappe.whitelist()
def export_management_kpis(days: int | str = 30):
	"""Kennzahlen als Excel-Mappe.

	Drei Blätter statt eines: Abteilungen, Mitarbeiter, Verlauf. In einem
	Blatt untereinander wären die Tabellen zwar auch lesbar, aber nicht mehr
	sortier- oder filterbar -- und genau dafür holt man sich die Zahlen nach
	Excel.
	"""
	_require_management()

	from io import BytesIO

	from openpyxl import Workbook
	from openpyxl.styles import Font
	from openpyxl.utils import get_column_letter

	data = build_kpis(days)
	book = Workbook()
	book.remove(book.active)

	def add_sheet(title: str, headers: list[str], rows: list[list]):
		sheet = book.create_sheet(title[:31])
		sheet.append(headers)
		for cell in sheet[1]:
			cell.font = Font(bold=True)
		for row in rows:
			sheet.append(row)

		# Spaltenbreite nach Inhalt -- sonst steht überall "####".
		for index, header in enumerate(headers, start=1):
			longest = max(
				[len(str(header))] + [len(str(row[index - 1])) for row in rows] or [0]
			)
			sheet.column_dimensions[get_column_letter(index)].width = min(longest + 3, 42)

		sheet.freeze_panes = "A2"

	add_sheet(
		_("Departments"),
		[
			_("Department"), _("Tickets"), _("Open"), _("Resolved"),
			_("Internal"), _("External"), _("External share %"),
			_("Avg response (h)"), _("Avg resolution (h)"),
			_("Hours total"), _("Hours billable"), _("Hours internal"),
			_("Hours submitted"), _("Hours draft"),
		],
		[
			[
				d["label"], d["total"], d["open"], d["closed"],
				d["internal"], d["external"], d["external_share"],
				d["avg_response_hours"], d["avg_resolution_hours"],
				d["hours_total"], d["hours_billable"], d["hours_internal"],
				d["hours_submitted"], d["hours_draft"],
			]
			for d in data["departments"]
		],
	)

	add_sheet(
		_("Employees"),
		[_("Employee"), _("Department"), _("Open tickets"), _("Hours"), _("Hours submitted")],
		[
			[e["employee_name"], e["department"], e["open_tickets"], e["hours"], e["hours_submitted"]]
			for e in data["employees"]
		],
	)

	departments = [d["label"] for d in data["departments"]]
	add_sheet(
		_("Trend"),
		[_("Period")] + departments + [_("Total")],
		[
			[point["period"]] + [point["values"].get(label, 0) for label in departments] + [point["total"]]
			for point in data["trend"]
		],
	)

	stream = BytesIO()
	book.save(stream)

	frappe.response["type"] = "binary"
	frappe.response["filecontent"] = stream.getvalue()
	frappe.response["filename"] = f"ticket-billing-kpi-{data['days']}d-{nowdate()}.xlsx"


# ---------------------------------------------------------------------------
# Abteilungssicht
# ---------------------------------------------------------------------------


def _resolve_department(department: str | None) -> str:
	"""Zuständige Abteilung bestimmen und Zugriff prüfen."""
	if not department:
		scope = get_scope_departments()
		department = scope[0] if scope else None

	if not department:
		frappe.throw(_("No department could be determined."), title=_("Department missing"))

	if not is_unrestricted():
		if not is_lead() or department not in get_scope_departments():
			frappe.throw(
				_("You are not allowed to access department {0}.").format(department),
				frappe.PermissionError,
			)

	return department


def _employee_trend(department: str, since: str, daily: bool) -> list[dict]:
	"""Aufkommen je Mitarbeiter im Zeitverlauf.

	Gezählt wird nach zugewiesenem Bearbeiter. Unzugewiesene Tickets fallen
	damit heraus -- sie stehen als eigene Kennzahl in der Freigabesicht und
	würden hier eine Linie ohne Namen erzeugen.
	"""
	bucket = "date(i.creation)" if daily else "date_format(i.creation, '%%x-KW%%v')"

	rows = frappe.db.sql(
		f"""
		select {bucket} as period, e.employee_name as who, count(i.name) as count
		from `tabIssue` i
		inner join `tabEmployee` e on e.name = i.`{FIELD_ASSIGNEE}`
		where i.creation >= %(since)s and i.`{FIELD_DEPARTMENT}` = %(department)s
		group by period, who
		order by min(i.creation) asc
		""",
		{"since": since, "department": department},
		as_dict=True,
	)

	counted: dict[str, dict] = {}
	for row in rows:
		key = str(row.period)
		entry = counted.setdefault(key, {"period": key, "values": {}, "total": 0})
		entry["values"][row.who] = row.count
		entry["total"] += row.count

	return [
		counted.get(key, {"period": key, "values": {}, "total": 0})
		for key in _period_keys(since, daily)
	]


def _pending_time(department: str) -> dict:
	"""Noch nicht gebuchte Zeiterfassungen der Abteilung.

	Ohne Zeitraumfilter: Ein Entwurf von vor drei Monaten ist genau der Fall,
	den man sehen will -- ihn aus dem gewählten Zeitraum herauszufiltern
	verstecke ihn ausgerechnet dann, wenn er am dringendsten ist.
	"""
	row = frappe.db.sql(
		"""
		select count(name) as entries, coalesce(sum(total_hours), 0) as hours,
		       min(creation) as oldest
		from `tabTimesheet`
		where docstatus = 0 and department = %(department)s
		""",
		{"department": department},
		as_dict=True,
	)[0]

	return {
		"entries": int(row.entries or 0),
		"hours": flt(row.hours, 2),
		"oldest": row.oldest,
	}


@frappe.whitelist()
def get_department_kpis(department: str | None = None, days: int | str = 84):
	"""Kennzahlen einer einzelnen Abteilung -- für die Leitung."""
	department = _resolve_department(department)
	days, since = _period(days)
	daily = days <= 31

	tickets = _department_tickets(since).get(department)
	hours = _department_hours(since).get(department)

	def count(field):
		return int(flt(tickets.get(field))) if tickets else 0

	# Auslastung: nur das eigene Team, in derselben Form wie beim
	# Gesamtdashboard, damit beide Ansichten dieselbe Komponente nutzen.
	employees = [e for e in _employees(since) if e["department"] == _short(department)]

	total = count("total")
	external = count("external_count")

	return {
		"department": department,
		"label": _short(department),
		"days": days,
		"since": since,
		"trend": _employee_trend(department, since, daily),
		"trend_granularity": "day" if daily else "week",
		"employees": employees,
		"pending_time": _pending_time(department),
		"totals": {
			"tickets": total,
			"open": count("open_count"),
			"closed": count("closed_count"),
			"internal": count("internal_count"),
			"external": external,
			"external_share": flt(external / total * 100, 1) if total else 0,
			"avg_response_hours": flt(tickets.avg_response_hours, 2)
			if tickets and tickets.avg_response_hours
			else None,
			"avg_resolution_hours": flt(tickets.avg_resolution_hours, 2)
			if tickets and tickets.avg_resolution_hours
			else None,
			"hours_total": flt(hours.hours_total if hours else 0, 2),
			"hours_billable": flt(hours.hours_billable if hours else 0, 2),
			"hours_internal": flt(hours.hours_internal if hours else 0, 2),
			"hours_submitted": flt(hours.hours_submitted if hours else 0, 2),
			"hours_draft": flt(hours.hours_draft if hours else 0, 2),
		},
	}
